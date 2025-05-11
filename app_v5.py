import streamlit as st
import os
import io
from PIL import Image
import ssl
from typing import Dict, Optional, Union
import aiohttp
import certifi
import asyncio
import base64
from pathlib import Path
import pandas as pd
import fitz
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def convert_pdf_to_images(pdf_bytes):
    images = []
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    for page in doc:
        pix = page.get_pixmap()
        img = Image.open(io.BytesIO(pix.tobytes("png")))
        images.append(img)
    return images


def generar_excel(datos_factura, total_tokens):
    output = io.BytesIO()
    # It's good practice to ensure 'openpyxl' is installed in the environment.
    # e.g., by adding 'openpyxl' to your requirements.txt
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        workbook = writer.book

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"
        )
        alignment_center = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        alignment_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border_side = Side(border_style="thin", color="000000")
        thin_border = Border(
            left=thin_border_side,
            right=thin_border_side,
            top=thin_border_side,
            bottom=thin_border_side,
        )

        def apply_styles_and_adjust_cols(sheet_name, df):
            sheet = workbook[sheet_name]
            # Aplicar estilo a los encabezados
            for col_num, value in enumerate(df.columns.values, 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment_center
                cell.border = thin_border

            # Aplicar estilo a las celdas de datos y ajustar ancho de columnas
            for row_num, row_data in enumerate(df.values, 2):
                for col_num, cell_value in enumerate(row_data, 1):
                    cell = sheet.cell(row=row_num, column=col_num)
                    cell.alignment = (
                        alignment_left
                        if isinstance(cell_value, str)
                        else alignment_center
                    )
                    cell.border = thin_border

            # Ajustar ancho de columnas
            for col_idx, column in enumerate(sheet.columns, 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column_letter].width = adjusted_width

            # Estilo especial para la hoja General para diferenciar secciones
            if sheet_name == "General":
                current_section = None
                for row_num in range(2, sheet.max_row + 1):
                    section_cell = sheet.cell(row=row_num, column=1)
                    if section_cell.value != current_section:
                        current_section = section_cell.value
                        # Aplicar un relleno diferente o un borde superior más grueso para la primera fila de una nueva sección
                        for col_num in range(1, sheet.max_column + 1):
                            cell = sheet.cell(row=row_num, column=col_num)
                            # Ejemplo: Poner negrita la primera celda de la sección
                            if col_num == 1:
                                cell.font = Font(bold=True)
                            # Podrías añadir un borde superior más grueso aquí si es la primera fila de la sección
                            # if row_num > 2: # Evitar aplicar al primer dato después del encabezado si no es una nueva sección
                            #     top_border = Border(top=Side(border_style="medium", color="000000"))
                            #     cell.border = Border(left=cell.border.left, right=cell.border.right, top=top_border, bottom=cell.border.bottom)

        # Sheet: General
        general_data_list = []
        if "emisor_receptor" in datos_factura:
            er_data = datos_factura.get("emisor_receptor", {})
            if er_data:  # Check if not empty
                comprobante = er_data.get("comprobante", {})
                for key, value in comprobante.items():
                    general_data_list.append(
                        {"Sección": "Comprobante", "Campo": key, "Valor": value}
                    )

                emisor = er_data.get("emisor", {})
                for key, value in emisor.items():
                    general_data_list.append(
                        {"Sección": "Emisor", "Campo": key, "Valor": value}
                    )

                receptor = er_data.get("receptor", {})
                for key, value in receptor.items():
                    general_data_list.append(
                        {"Sección": "Receptor", "Campo": key, "Valor": value}
                    )

                otros = er_data.get("otros", {})
                for key, value in otros.items():
                    general_data_list.append(
                        {"Sección": "Otros", "Campo": key, "Valor": value}
                    )

        if general_data_list:
            df_general = pd.DataFrame(general_data_list)
            df_general.to_excel(writer, sheet_name="General", index=False)
            apply_styles_and_adjust_cols("General", df_general)
        else:
            df_no_data = pd.DataFrame(
                [{"Mensaje": "No hay datos generales disponibles."}]
            )
            df_no_data.to_excel(writer, sheet_name="General", index=False)
            apply_styles_and_adjust_cols("General", df_no_data)

        # Sheet: Detalle Factura (Items)
        items_list = []
        if "items" in datos_factura and datos_factura.get("items", {}).get("detalles"):
            items_list = datos_factura["items"]["detalles"]

        if items_list:
            df_items = pd.DataFrame(items_list)
            df_items.to_excel(writer, sheet_name="Detalle Factura", index=False)
            apply_styles_and_adjust_cols("Detalle Factura", df_items)
        else:
            df_no_data = pd.DataFrame(
                [{"Mensaje": "No hay detalles de items disponibles."}]
            )
            df_no_data.to_excel(writer, sheet_name="Detalle Factura", index=False)
            apply_styles_and_adjust_cols("Detalle Factura", df_no_data)

        # Sheet: Totales y Observaciones
        summary_data_list = []
        if "items" in datos_factura:
            items_data = datos_factura.get("items", {})
            if "subtotal" in items_data:  # Check key existence
                summary_data_list.append(
                    {"Descripción": "Subtotal", "Valor": items_data.get("subtotal")}
                )
            if "total" in items_data:  # Check key existence
                summary_data_list.append(
                    {"Descripción": "Total", "Valor": items_data.get("total")}
                )
            if "observaciones" in items_data and items_data.get("observaciones"):
                summary_data_list.append(
                    {
                        "Descripción": "Observaciones",
                        "Valor": items_data.get("observaciones"),
                    }
                )

        if summary_data_list:
            df_summary = pd.DataFrame(summary_data_list)
            df_summary.to_excel(
                writer, sheet_name="Totales y Observaciones", index=False
            )
            apply_styles_and_adjust_cols("Totales y Observaciones", df_summary)
        else:
            df_no_data = pd.DataFrame(
                [{"Mensaje": "No hay totales ni observaciones disponibles."}]
            )
            df_no_data.to_excel(
                writer, sheet_name="Totales y Observaciones", index=False
            )
            apply_styles_and_adjust_cols("Totales y Observaciones", df_no_data)

        # Sheet: Impuestos
        impuestos_list = []
        if "impuestos" in datos_factura and datos_factura.get("impuestos", {}).get(
            "impuestos"
        ):
            impuestos_list = datos_factura["impuestos"]["impuestos"]

        if impuestos_list:
            df_impuestos = pd.DataFrame(impuestos_list)
            df_impuestos.to_excel(writer, sheet_name="Impuestos", index=False)
            apply_styles_and_adjust_cols("Impuestos", df_impuestos)
        else:
            df_no_data = pd.DataFrame([{"Mensaje": "No hay impuestos registrados."}])
            df_no_data.to_excel(writer, sheet_name="Impuestos", index=False)
            apply_styles_and_adjust_cols("Impuestos", df_no_data)

        # Sheet: Retenciones
        retenciones_list = []
        if "impuestos" in datos_factura and datos_factura.get("impuestos", {}).get(
            "retenciones"
        ):
            retenciones_list = datos_factura["impuestos"]["retenciones"]

        if retenciones_list:
            df_retenciones = pd.DataFrame(retenciones_list)
            df_retenciones.to_excel(writer, sheet_name="Retenciones", index=False)
            apply_styles_and_adjust_cols("Retenciones", df_retenciones)
        else:
            df_no_data = pd.DataFrame([{"Mensaje": "No hay retenciones registradas."}])
            df_no_data.to_excel(writer, sheet_name="Retenciones", index=False)
            apply_styles_and_adjust_cols("Retenciones", df_no_data)

        # Sheet: Uso de Tokens
        if total_tokens:
            tokens_list_excel = [
                {"Tipo": key, "Cantidad": value} for key, value in total_tokens.items()
            ]
            df_tokens = pd.DataFrame(tokens_list_excel)
            df_tokens.to_excel(writer, sheet_name="Uso de Tokens", index=False)
            apply_styles_and_adjust_cols("Uso de Tokens", df_tokens)
        else:
            df_no_data = pd.DataFrame(
                [{"Mensaje": "No hay datos de uso de tokens disponibles."}]
            )
            df_no_data.to_excel(writer, sheet_name="Uso de Tokens", index=False)
            apply_styles_and_adjust_cols("Uso de Tokens", df_no_data)

    output.seek(0)
    return output


# Mostrar datos extraídos
def mostrar_datos(respuestas):
    # Extracción de datos de las respuestas
    datos_factura = {}
    total_tokens = {
        "input_tokens": 0,
        "output_tokens": 0,
        "cache_creation_input_tokens": 0,
        "cache_read_input_tokens": 0,
    }

    for respuesta in respuestas:
        # Acumular tokens
        for token_type, value in respuesta.get("usage", {}).items():
            total_tokens[token_type] += value

        # Extraer datos según el tipo de herramienta
        if respuesta.get("content") and len(respuesta["content"]) > 0:
            tool_content = respuesta["content"][0]
            if tool_content.get("name") == "datos_del_emisor_y_receptor":
                datos_factura["emisor_receptor"] = tool_content.get("input", {})
            elif tool_content.get("name") == "detalle_de_items_facturados":
                datos_factura["items"] = tool_content.get("input", {})
            elif tool_content.get("name") == "impuestos_y_retenciones_de_la_factura":
                datos_factura["impuestos"] = tool_content.get("input", {})

    # Mostrar datos del comprobante, emisor y receptor
    if "emisor_receptor" in datos_factura:
        mostrar_datos_comprobante(datos_factura["emisor_receptor"])

    # Mostrar detalles de facturación
    if "items" in datos_factura:
        mostrar_items_facturados(datos_factura["items"])

    # Mostrar impuestos y retenciones
    if "impuestos" in datos_factura:
        mostrar_impuestos(datos_factura["impuestos"])

    # Mostrar el uso de tokens
    mostrar_uso_tokens(total_tokens)

    # Generar Excel y agregar botón de descarga
    if datos_factura or total_tokens:  # Check if there's any data to put in Excel
        try:
            excel_bytes = generar_excel(datos_factura, total_tokens)
            st.download_button(
                label="Descargar Factura en Excel",
                data=excel_bytes,
                file_name="factura_procesada.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            st.error(f"Error al generar el archivo Excel: {e}")
            # Optionally log the full traceback for debugging
            # import traceback
            # st.expander("Detalles del error").code(traceback.format_exc())


# Mostrar datos del comprobante
def mostrar_datos_comprobante(datos):
    st.header("Datos del Comprobante")

    # Comprobante
    if "comprobante" in datos:
        comprobante = datos["comprobante"]
        st.subheader("Comprobante")

        # Crear dos columnas para los datos del comprobante
        col1, col2 = st.columns(2)

        with col1:
            st.info(f"**Tipo:** {comprobante.get('tipo', 'N/A')}")
            st.info(f"**Número:** {comprobante.get('numero', 'N/A')}")
            st.info(f"**Moneda:** {comprobante.get('moneda', 'N/A')}")

        with col2:
            st.info(f"**Fecha de emisión:** {comprobante.get('fecha_emision', 'N/A')}")
            st.info(
                f"**Jurisdiccion fiscal:** {comprobante.get('jurisdiccion_fiscal', 'N/A')}"
            )
            if "subtipo" in comprobante:
                st.info(f"**Subtipo:** {comprobante.get('subtipo', 'N/A')}")

    # Emisor
    if "emisor" in datos:
        emisor = datos["emisor"]
        st.subheader("Datos del Emisor")

        # Mostrar datos del emisor sin columnas anidadas
        st.info(f"**Nombre:** {emisor.get('nombre', 'N/A')}")
        st.info(f"**ID Fiscal:** {emisor.get('id_fiscal', 'N/A')}")
        if "condicion_iva" in emisor:
            st.info(f"**Condición IVA:** {emisor.get('condicion_iva', 'N/A')}")
        if "direccion" in emisor:
            st.info(f"**Dirección:** {emisor.get('direccion', 'N/A')}")

    # Receptor
    if "receptor" in datos:
        receptor = datos["receptor"]
        st.subheader("Datos del Receptor")

        # Mostrar datos del receptor sin columnas anidadas
        st.info(f"**Nombre:** {receptor.get('nombre', 'N/A')}")
        if "id_fiscal" in receptor:
            st.info(f"**ID Fiscal:** {receptor.get('id_fiscal', 'N/A')}")
        if "condicion_iva" in receptor:
            st.info(f"**Condición IVA:** {receptor.get('condicion_iva', 'N/A')}")
        if "direccion" in receptor:
            st.info(f"**Dirección:** {receptor.get('direccion', 'N/A')}")

    # Otros datos
    if "otros" in datos:
        otros = datos["otros"]
        st.subheader("Información Adicional")

        # Mostrar datos adicionales sin columnas anidadas
        if "forma_pago" in otros:
            st.info(f"**Forma de pago:** {otros.get('forma_pago', 'N/A')}")
        if "CAE" in otros:
            st.info(f"**CAE:** {otros.get('CAE', 'N/A')}")
        if "vencimiento_CAE" in otros:
            st.info(f"**Vencimiento CAE:** {otros.get('vencimiento_CAE', 'N/A')}")


# Mostrar detalles de facturación
def mostrar_items_facturados(datos):
    st.header("Detalles de Facturación")

    # Tabla de detalles
    if (
        "detalles" in datos
        and isinstance(datos["detalles"], list)
        and datos["detalles"]
    ):
        st.subheader("Ítems Facturados")
        try:
            df = pd.DataFrame(datos["detalles"])

            # Formatear datos numéricos
            if "precio_unitario" in df:
                df["precio_unitario"] = df["precio_unitario"].apply(
                    lambda x: f"${x:,.2f}" if pd.notna(x) else "N/A"
                )
            if "precio_total" in df:
                df["precio_total"] = df["precio_total"].apply(
                    lambda x: f"${x:,.2f}" if pd.notna(x) else "N/A"
                )

            st.dataframe(df, use_container_width=True)
        except ValueError as e:
            with st.expander(
                "No fue posible formatear correctamente los detalles de facturación. A continuación, se muestra el contenido en formato JSON:"
            ):
                st.write(datos["detalles"])
    else:
        with st.expander(
            "No fue posible formatear correctamente los detalles de facturación. A continuación, se muestra el contenido en formato JSON:"
        ):
            st.write(datos["detalles"])

    # Totales
    st.subheader("Totales")
    if "subtotal" in datos and pd.notna(datos["subtotal"]):
        try:
            # Convert subtotal to float if it's a string
            subtotal = (
                float(datos["subtotal"])
                if isinstance(datos["subtotal"], str)
                else datos["subtotal"]
            )
            st.success(f"**Subtotal:** ${subtotal:,.2f}")
        except (ValueError, TypeError):
            # Handle case where conversion fails
            st.success(f"**Subtotal:** ${datos['subtotal']}")
    else:
        st.info("Subtotal no disponible.")

    if "total" in datos and pd.notna(datos["total"]):
        try:
            # Convert total to float if it's a string
            total = (
                float(datos["total"])
                if isinstance(datos["total"], str)
                else datos["total"]
            )
            st.success(f"**Total:** ${total:,.2f}")
        except (ValueError, TypeError):
            # Handle case where conversion fails
            st.success(f"**Total:** ${datos['total']}")
    else:
        st.info("Total no disponible.")

    # Observaciones
    if "observaciones" in datos and datos["observaciones"]:
        st.subheader("Observaciones")
        st.info(datos["observaciones"])


def mostrar_impuestos(datos):
    st.header("Impuestos y Retenciones")

    # Tabla de impuestos
    if "impuestos" in datos and datos["impuestos"]:
        st.subheader("Impuestos")

        # Crear una tabla para mostrar los impuestos
        datos_impuestos = []
        for impuesto in datos["impuestos"]:
            tipo = impuesto.get("tipo", "N/A")
            if "descripcion" in impuesto and impuesto["descripcion"]:
                tipo += f" - {impuesto['descripcion']}"

            alicuota = (
                f"{impuesto.get('alicuota', 'N/A')}%"
                if "alicuota" in impuesto and impuesto["alicuota"] is not None
                else "N/A"
            )
            importe = f"${impuesto.get('importe', 0):,.2f}"

            datos_impuestos.append(
                {"Tipo": tipo, "Alícuota": alicuota, "Importe": importe}
            )

        # Mostrar como dataframe si hay datos
        if datos_impuestos:
            st.dataframe(
                pd.DataFrame(datos_impuestos), use_container_width=True, hide_index=True
            )
    else:
        st.info("No hay impuestos registrados")

    # Tabla de retenciones
    if "retenciones" in datos and datos["retenciones"]:
        st.subheader("Retenciones")

        # Crear una tabla para mostrar las retenciones
        datos_retenciones = []
        for retencion in datos["retenciones"]:
            tipo = retencion.get("tipo", "N/A")
            if "descripcion" in retencion and retencion["descripcion"]:
                tipo += f" - {retencion['descripcion']}"

            base = f"${retencion.get('base_imponible', 0):,.2f}"

            datos_retenciones.append({"Tipo": tipo, "Base Imponible": base})

        # Mostrar como dataframe si hay datos
        if datos_retenciones:
            st.dataframe(
                pd.DataFrame(datos_retenciones),
                use_container_width=True,
                hide_index=True,
            )
    else:
        st.info("No hay retenciones registradas")


# Mostrar uso de tokens
def mostrar_uso_tokens(tokens):
    st.header("Uso de Tokens")

    # Crear una tabla para el uso de tokens
    tokens_data = {
        "Tipo": [
            "Input Tokens",
            "Output Tokens",
            "Cache Creation",
            "Cache Read",
        ],
        "Cantidad": [
            tokens["input_tokens"],
            tokens["output_tokens"],
            tokens["cache_creation_input_tokens"],
            tokens["cache_read_input_tokens"],
        ],
    }

    # Mostrar como dataframe
    df_tokens = pd.DataFrame(tokens_data)
    st.dataframe(df_tokens, use_container_width=True, hide_index=True)

    # Mostrar el total como un métrico destacado
    # st.metric("Total Tokens Utilizados", sum(tokens.values()))


def convert_image_to_base64(img: Image.Image) -> str:
    """
    Converts a PIL Image to a base64-encoded string.

    Args:
    img (Image.Image): The PIL image to be converted.

    Returns:
    str: The base64-encoded string representation of the image.
    """
    buffered = io.BytesIO()
    img.save(buffered, format="PNG")
    base64_encoded_data = base64.b64encode(buffered.getvalue())
    base64_string = base64_encoded_data.decode("utf-8")
    return base64_string


def pdf_to_base64(file_path: str) -> Union[str, None]:
    try:
        with open(file_path, "rb") as pdf_file:
            binary_data = pdf_file.read()
            base_64_encoded_data = base64.b64encode(binary_data)
            base64_string = base_64_encoded_data.decode("utf-8")
        return base64_string
    except Exception as e:
        print(f"An error occurred: {e}")
        return None


async def make_api_request(
    url: str, headers: Dict, data: Dict, process_id: str, retries: int = 5
) -> Optional[Dict]:
    """Alternative helper function to make API requests with retries asynchronously."""
    # Create an SSL context using the certifi CA bundle
    ssl_context = ssl.create_default_context(cafile=certifi.where())
    # Create a connector with the custom SSL context
    connector = aiohttp.TCPConnector(ssl=ssl_context)
    connector = aiohttp.TCPConnector(verify_ssl=False)
    async with aiohttp.ClientSession(connector=connector) as session:
        for i in range(retries):
            try:
                async with session.post(url, headers=headers, json=data) as response:
                    if response.status == 200:
                        return await response.json()
                    elif response.status in [429, 529, 503]:
                        retry_after = response.headers.get("retry-after")
                        print(f"retry_after: {retry_after}")
                        print(
                            f"Service unavailable. Waiting {retry_after} seconds before retrying..."
                        )
                        await asyncio.sleep(retry_after)
                        # print(
                        #     f"Service unavailable. Waiting {WAIT_TIMES[i]} seconds before retrying..."
                        # )
                        # await asyncio.sleep(WAIT_TIMES[i])
                    else:
                        print(f"Error: {response.status} - {await response.text()}")
                        raise ValueError(
                            f"Request failed with status {response.status}"
                        )
            except aiohttp.ClientError as e:
                raise ValueError(f"Request error: {str(e)}")
    raise ValueError("Max retries exceeded.")


async def call_claude_vision(
    tools: list,
    encoded_img: str,
    type: str,
    prompt: str,
    tool_name: str,
    process_id: str,
    model: str = "claude-3-7-sonnet-20250219",
):
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if api_key is None:
        raise ValueError("ANTHROPIC_API_KEY is not set in the environment variables.")
    data = {
        "model": model,
        "tools": tools,
        "max_tokens": 8192,
        "messages": [
            {
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": type,
                            "data": encoded_img,
                        },
                        "cache_control": {"type": "ephemeral"},
                    },
                    {"type": "text", "text": prompt},
                ],
            },
        ],
        "tool_choice": {
            "type": "tool",
            "name": tool_name,
            "disable_parallel_tool_use": True,
        },
    }

    headers = {
        "Content-Type": "application/json",
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01",
    }

    response = await make_api_request(
        url="https://api.anthropic.com/v1/messages",
        headers=headers,
        data=data,
        process_id=process_id,
    )
    return response


async def call_claude_pdf(
    tools: list,
    static_content: str,
    prompt: str,
    tool_name: str,
    process_id: str,
    model: str = "claude-3-7-sonnet-20250219",
):
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if api_key is None:
        raise ValueError("ANTHROPIC_API_KEY is not set in the environment variables.")

    data = {
        "model": model,
        "tools": tools,
        "max_tokens": 8192,
        "messages": [
            {
                "role": "user",
                "content": [
                    {
                        "type": "document",
                        "source": {
                            "type": "base64",
                            "media_type": "application/pdf",
                            "data": static_content,
                        },
                        "cache_control": {"type": "ephemeral"},
                    },
                    {"type": "text", "text": prompt},
                ],
            },
        ],
        "tool_choice": {
            "type": "tool",
            "name": tool_name,
            "disable_parallel_tool_use": True,
        },
    }

    headers = {
        "Content-Type": "application/json",
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01",
    }

    response = await make_api_request(
        url="https://api.anthropic.com/v1/messages",
        headers=headers,
        data=data,
        process_id=process_id,
    )
    return response


def save_uploaded_file(uploaded_file, save_path):
    # Create the full file path
    file_path = os.path.join(save_path, uploaded_file.name)

    # Ensure the directory exists
    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    # Save the file
    with open(file_path, "wb") as f:
        f.write(uploaded_file.getbuffer())
    return file_path


async def main():
    # Configuración de la página
    st.set_page_config(page_title="Ticket AI", page_icon="📄", layout="wide")

    # Título y descripción
    st.sidebar.title("Ticket AI 📄")
    st.sidebar.write(
        "Ticket AI es una aplicación diseñada para extraer información relevante de facturas."
    )

    # Separador
    st.sidebar.markdown("---")

    # Sección de carga de archivos
    st.sidebar.subheader("Cargar Factura")
    st.sidebar.write(
        "Por favor, cargue su factura en formato PDF o imagen (PNG, JPEG, WEBP, GIF no animado)."
    )

    # Componente para cargar archivos
    archivo_cargado = st.sidebar.file_uploader(
        "Seleccione un archivo",
        type=["pdf", "png", "jpg", "jpeg", "webp", "gif"],
        help="Formatos soportados: PDF, PNG, JPEG, WEBP, GIF no animado",
    )

    # Mostrar el archivo cargado
    if archivo_cargado is not None:
        if st.sidebar.button("Procesar Factura", use_container_width=True):
            tools = [
                {
                    "prompt": (
                        "Extrae los detalles estructurados de la factura. "
                        "Incluye el domicilio comercial del emisor y del receptor, "
                        "priorizando siempre el Domicilio Comercial sobre el Legal o Fiscal. "
                    ),
                    "data": {
                        "name": "datos_del_emisor_y_receptor",  # Updated name
                        "description": "Extrae y gestiona los datos clave de un comprobante fiscal, incluyendo datos del emisor y del receptor con validación de enums.",
                        "input_schema": {
                            "type": "object",
                            "properties": {
                                "comprobante": {
                                    "type": "object",
                                    "description": "Datos generales del comprobante.",
                                    "properties": {
                                        "tipo": {
                                            "type": "string",
                                            "enum": [
                                                "Factura",
                                                "Nota de Crédito",
                                                "Nota de Débito",
                                            ],
                                            "description": "Tipo de comprobante fiscal.",
                                        },
                                        "subtipo": {
                                            "type": "string",
                                            "enum": [
                                                "Para operaciones entre responsables inscriptos",
                                                "Para consumidores finales y exentos",
                                                "Emitida por monotributistas",
                                            ],
                                            "description": "Subtipo del comprobante basado en el régimen fiscal del receptor.",
                                        },
                                        "jurisdiccion_fiscal": {
                                            "type": "string",
                                            "description": "País o jurisdicción fiscal del comprobante, como 'Argentina'.",
                                        },
                                        "numero": {
                                            "type": "string",
                                            "description": "Número del comprobante, por ejemplo: '0001-00001234'.",
                                        },
                                        "fecha_emision": {
                                            "type": "string",
                                            "format": "date",
                                            "description": "Fecha de emisión del comprobante (YYYY-MM-DD).",
                                        },
                                        "moneda": {
                                            "type": "string",
                                            "description": "Moneda en la que se emite el comprobante, por ejemplo 'ARS', 'USD'.",
                                        },
                                    },
                                    "required": [
                                        "tipo",
                                        "jurisdiccion_fiscal",
                                        "numero",
                                        "fecha_emision",
                                        "moneda",
                                    ],
                                    "additionalProperties": False,
                                },
                                "emisor": {
                                    "type": "object",
                                    "description": "Datos del emisor del comprobante.",
                                    "properties": {
                                        "nombre": {
                                            "type": "string",
                                            "description": "Nombre o razón social del emisor.",
                                        },
                                        "id_fiscal": {
                                            "type": "string",
                                            "description": "CUIT o CUIL del emisor.",
                                        },
                                        "direccion": {
                                            "type": "string",
                                            "description": "Extrae el domicilio comercial del emisor. Prioriza siempre el Domicilio Comercial sobre el Legal o Fiscal.",
                                        },
                                        "condicion_iva": {
                                            "type": "string",
                                            "enum": [
                                                "Responsable Inscripto",
                                                "Monotributo",
                                                "Exento",
                                                "Desconocido",
                                            ],
                                            "description": "Condición frente al IVA del emisor.",
                                        },
                                    },
                                    "required": ["nombre", "id_fiscal"],
                                    "additionalProperties": False,
                                },
                                "receptor": {
                                    "type": "object",
                                    "description": "Datos del receptor del comprobante.",
                                    "properties": {
                                        "nombre": {
                                            "type": "string",
                                            "description": "Nombre o razón social del receptor.",
                                        },
                                        "id_fiscal": {
                                            "type": "string",
                                            "description": "CUIT o CUIL del receptor.",
                                        },
                                        "direccion": {
                                            "type": "string",
                                            "description": "Extrae el domicilio comercial del receptor. Prioriza siempre el Domicilio Comercial sobre el Legal o Fiscal.",
                                        },
                                        "condicion_iva": {
                                            "type": "string",
                                            "enum": [
                                                "Responsable Inscripto",
                                                "Consumidor Final",
                                                "Desconocido",
                                            ],
                                            "description": "Condición frente al IVA del receptor.",
                                        },
                                    },
                                    "required": ["nombre"],
                                    "additionalProperties": False,
                                },
                                "otros": {
                                    "type": "object",
                                    "description": "Otros datos relevantes del comprobante.",
                                    "properties": {
                                        "forma_pago": {
                                            "type": "string",
                                            "description": "Método de pago utilizado. Ejemplos: 'Efectivo', 'Transferencia', 'Tarjeta de crédito'.",
                                            "nullable": True,
                                        },
                                        "CAE": {
                                            "type": "string",
                                            "description": "Código de Autorización Electrónico emitido por AFIP.",
                                            "nullable": True,
                                        },
                                        "vencimiento_CAE": {
                                            "type": "string",
                                            "format": "date",
                                            "description": "Fecha de vencimiento del CAE en formato YYYY-MM-DD.",
                                            "nullable": True,
                                        },
                                        "additionalProperties": False,
                                    },
                                },
                            },
                            "required": ["comprobante", "emisor", "receptor"],
                            "additionalProperties": False,
                        },
                    },
                },
                {
                    "prompt": "Extrae los detalles estructurados de la factura",
                    "data": {
                        "name": "detalle_de_items_facturados",  # Updated name
                        "description": "Extrae y valida los detalles de ítems facturados junto con totales e impuestos.",
                        "input_schema": {
                            "type": "object",
                            "properties": {
                                "detalles": {
                                    "type": "array",
                                    "description": "Lista de ítems facturados en la factura.",
                                    "items": {
                                        "type": "object",
                                        "properties": {
                                            "descripcion": {
                                                "type": "string",
                                                "description": "Descripción del producto o servicio facturado.",
                                            },
                                            "cantidad": {
                                                "type": "number",
                                                "description": "Cantidad facturada.",
                                            },
                                            "precio_unitario": {
                                                "type": "number",
                                                "description": "Precio por unidad antes de impuestos.",
                                            },
                                            "precio_total": {
                                                "type": "number",
                                                "description": "Total de la línea (cantidad x precio unitario).",
                                            },
                                        },
                                        "required": [
                                            "descripcion",
                                            "cantidad",
                                            "precio_unitario",
                                            "precio_total",
                                        ],
                                        "additionalProperties": False,
                                    },
                                },
                                "subtotal": {
                                    "type": "number",
                                    "description": "Suma de importes sin incluir impuestos.",
                                },
                                "total": {
                                    "type": "number",
                                    "description": "Importe total del comprobante, calculado como subtotal + impuestos - retenciones. Ejemplo: 4130.00",
                                },
                                "observaciones": {
                                    "type": "string",
                                    "description": "Notas o comentarios adicionales sobre el comprobante.",
                                },
                            },
                            "required": ["detalles", "subtotal", "total"],
                            "additionalProperties": False,
                        },
                    },
                },
                {
                    "prompt": (
                        "Extrae y clasifica todos los cargos que aparezcan en la factura por encima del subtotal. "
                        "Incluye impuestos como IVA o impuestos a las ventas, y detalla cualquier cargo adicional o recargo con su descripción y monto. "
                        "Las retenciones deben corresponder exclusivamente a conceptos fiscales o tributarios (como Ganancias, IVA, IIBB). "
                        "No incluyas descuentos comerciales, promociones ni bonificaciones bajo la categoría de retenciones. "
                        "Estos deben clasificarse por separado o ser ignorados si no corresponden a un cargo sobre el subtotal."
                    ),
                    "data": {
                        "name": "impuestos_y_retenciones_de_la_factura",  # Updated name
                        "description": "Extrae, valida y estructura la información de impuestos y retenciones de una factura.",
                        "input_schema": {
                            "type": "object",
                            "properties": {
                                "impuestos": {
                                    "type": "array",
                                    "description": "Lista de impuestos aplicados en la factura.",
                                    "items": {
                                        "type": "object",
                                        "properties": {
                                            "tipo": {
                                                "type": "string",
                                                "description": "Tipo de impuesto aplicado (e.g., IVA, Percepción, Impuesto Municipal).",
                                            },
                                            "descripcion": {
                                                "type": "string",
                                                "description": "Descripción detallada del impuesto.",
                                                "nullable": True,
                                            },
                                            "base_imponible": {
                                                "type": "number",
                                                "description": "Monto base sobre el que se calcula el impuesto.",
                                            },
                                            "alicuota": {
                                                "type": "number",
                                                "description": "Porcentaje de alícuota aplicada.",
                                                "nullable": True,
                                            },
                                            "importe": {
                                                "type": "number",
                                                "description": "Importe del impuesto calculado.",
                                            },
                                        },
                                        "required": [
                                            "tipo",
                                            "base_imponible",
                                            "importe",
                                        ],
                                        "additionalProperties": False,
                                    },
                                },
                                "retenciones": {
                                    "type": "array",
                                    "description": "Lista de retenciones aplicadas. No incluir descuentos ni promociones; estos deben clasificarse por separado.",
                                    "items": {
                                        "type": "object",
                                        "properties": {
                                            "tipo": {
                                                "type": "string",
                                                "description": "Tipo de retención (e.g., Ganancias, IVA, IIBB). No usar para descuentos o promociones.",
                                            },
                                            "description": {
                                                "type": "string",
                                                "description": "Detalle adicional de la retención. No incluir información de descuentos ni promociones.",
                                                "nullable": True,
                                            },
                                            "base_imponible": {
                                                "type": "number",
                                                "description": "Monto base sobre el que se aplica la retención.",
                                            },
                                        },
                                        "required": ["tipo", "base_imponible"],
                                        "additionalProperties": False,
                                    },
                                },
                            },
                            "required": ["impuestos", "retenciones"],
                            "additionalProperties": False,
                        },
                    },
                },
            ]

            # Mostrar vista previa según el tipo de archivo
            if archivo_cargado.type.startswith("image"):
                col1, col2 = st.columns([1, 1])

                with col1:
                    st.write("## Factura")
                    st.image(
                        archivo_cargado,
                        caption=f"Imagen cargada: {archivo_cargado.name}",
                        use_container_width=True,
                    )

                file_path = save_uploaded_file(archivo_cargado, "./uploaded_pdfs")

                image_file = Path(file_path)
                assert image_file.is_file(), "The provided image path does not exist."

                # Read and encode the image file
                base64_string = base64.b64encode(image_file.read_bytes()).decode()

                response = await call_claude_vision(
                    tools=[tool["data"] for tool in tools],
                    encoded_img=base64_string,
                    type=archivo_cargado.type,
                    prompt=tools[0]["prompt"],
                    tool_name=tools[0]["data"]["name"],
                    process_id="process_id",
                )

                # with st.expander("Respuesta"):
                #     st.write(response)

                tasks = []
                for tool in tools[1:]:
                    tool_res = call_claude_vision(
                        tools=[tool["data"] for tool in tools],
                        encoded_img=base64_string,
                        type=archivo_cargado.type,
                        prompt=tool["prompt"],
                        tool_name=tool["data"]["name"],
                        process_id="process_id",
                    )
                    tasks.append(tool_res)

                # Wait for all tasks to complete
                results = await asyncio.gather(*tasks)
                # with st.expander("Respuestas"):
                #     st.write(results)

                all_results = [response] + results

                with col2:
                    st.write("## Respuestas")
                    with st.expander("Respuestas - No formateadas"):
                        st.write(all_results)
                    mostrar_datos(all_results)

                # Delete the temporary PDF file after processing
                try:
                    os.remove(file_path)
                except OSError as e:
                    print(f"Error deleting file {file_path}: {e}")

            elif archivo_cargado.type == "application/pdf":

                file_path = save_uploaded_file(archivo_cargado, "./uploaded_pdfs")
                static_content = pdf_to_base64(file_path)

                with open(file_path, "rb") as saved_file:
                    images = convert_pdf_to_images(saved_file.read())

                    col1, col2 = st.columns([1, 1])  # Puedes ajustar las proporciones

                    with col1:
                        st.write("## Páginas del PDF")
                        for i, img in enumerate(images):
                            st.image(img, caption=f"Página {i + 1}")

                    static_content = pdf_to_base64(file_path)

                    response = await call_claude_pdf(
                        tools=[tool["data"] for tool in tools],
                        static_content=static_content,
                        prompt=tools[0]["prompt"],
                        tool_name=tools[0]["data"]["name"],
                        process_id="process_id",
                    )

                    tasks = []
                    for tool in tools[1:]:
                        tool_res = call_claude_pdf(
                            tools=[tool["data"] for tool in tools],
                            static_content=static_content,
                            prompt=tool["prompt"],
                            tool_name=tool["data"]["name"],
                            process_id="process_id",
                        )
                        tasks.append(tool_res)

                    results = await asyncio.gather(*tasks)

                    all_results = [response] + results

                    with col2:
                        st.write("## Respuestas")
                        with st.expander("Respuestas - No formateadas"):
                            st.write(all_results)
                        mostrar_datos(all_results)

                try:
                    os.remove(file_path)
                except OSError as e:
                    print(f"Error deleting file {file_path}: {e}")


if __name__ == "__main__":
    asyncio.run(main())
